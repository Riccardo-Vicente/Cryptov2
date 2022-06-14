import pandas as pd
import numpy as np
from openpyxl.workbook import Workbook
from  openpyxl import load_workbook
#-------------------------------------------------------------------------------------------------------
#Renewable energy infrustructure
#------------------------------------------------------------------------------------------------------

# Defining Renewable energy costs ($/MWh)

LCOE = {
    'pv_energy_residential': 189,
    'pv_energy_largescale': 37,
    'wind_energy': 40,
}
renewable_type = 'pv_energy_largescale'

USD_Rand = 15

#Ask user renewable energy specs
RE_capacity = float(input("Enter RE capacity ")) # MW
print("Capacity: {} MW".format(RE_capacity))

#Determine cost of RE over lifetime
eff = 0.1724 # %
E_annual = RE_capacity*4380*eff
E_annual = float(E_annual)
print("Annual Energy generation: {} MWh.".format(E_annual))
Lifetime = 25 # years

E_lifetime = E_annual*Lifetime
print("Lifetime Energy generation: {} MWh".format(E_lifetime))

RE_cost = E_lifetime*LCOE[renewable_type]*USD_Rand
print("Capital Cost of RE infrustructure: R {:.2f}".format(RE_cost))
print("\n")
#------------------------------------------------------------------------------------------
#Crypto mining infrustructre
#------------------------------------------------------------------------------------------

# Get mining equipment data
#df = pd.read_excel('./Data/crypto_mining_equipment.xlsx')
wb = load_workbook('./Data/crypto_mining_equipment.xlsx')
ws = wb.active

# Declare each excel column to a list to refer to later.
miner_col = ws['C2':'C9']
print("Mining Hardware:")
miners = []
for cell in miner_col:
    for a in cell:
        miners.append(a.value)
print(miners)

power_col = ws['E2':'E9']
print("Power values:")
power = []
for cell in power_col:
    for a in cell:
        power.append(a.value)
print(power)

price_col = ws['D2':'D9']
print("Miner Prices:")
price = []
for cell in price_col:
    for a in cell:
        price.append(a.value)
print(price)

hash_rate_col = ws['F2':'F9']
print("Miner Hash rates:")
hash_rate = []
for cell in hash_rate_col:
    for a in cell:
        hash_rate.append(a.value)
print(hash_rate)

#Declare miners to purchase
miner1 = 'Canaan (Avalon 6)'
miner2 = 'GeForce RTX 3060 Ti'
print("\nSelected mining hardware:")

#Get selected miner specs
i = {
'ATI (Radeon 5850)': 0,
'ATI (Radeon 6990)': 1,
'Monarch BPU 600C': 2,
'Canaan (Avalon 6)': 3,
'Bitmain (Antminer S9)': 4,
'GeForce GTX 1080 Ti': 5,
'Bitmain (Antminer S19)': 6,
'GeForce RTX 3060 Ti': 7,
}
print("Miner 1: {}".format(miner1))
m1 = i[miner1]
print("Power Consumption: {} W".format(power[m1]))
print("Price: {} USD".format(price[m1]))
print("Hash Rate: {} GH/s".format(hash_rate[m1]))

print("\nMiner 2: {}".format(miner2))
m2 = i[miner2]
print("Power Consumption: {} W".format(power[m2]))
print("Price: {} USD".format(price[m2]))
print("Hash Rate: {} GH/s".format(hash_rate[m2]))






